import openpyxl
from pathlib import Path
from os import path
from email_send import send_emails

def reading_xlsx(filename):
    columns = {1:"A", 2:"B",3:"C",4:"D",5:"E",6:"F",7:"G",8:"H"}
    #print(path.abspath('.') + "\\" + filename) check path
    wb = openpyxl.load_workbook(path.abspath('.') + "\\" + filename)

    sheet = wb.active
    data = list()
    j = int(2)
    while sheet.cell(j,2).value != None:
        name = sheet.cell(j,2).value
        email = sheet.cell(j,3).value
        #number = sheet.cell(j,4).value
        res_school = sheet.cell(j, 5).value
        res_zno = sheet.cell(j, 6).value
        if name != None and email != None and res_school != None and res_zno != None:
            data.append({
                "name": name,
                "email": email,
                "res_school": res_school,
                "res_zno": res_zno,
            })
        else:
            pass
        j += 1

    send_emails(data)
    return(len(data))
